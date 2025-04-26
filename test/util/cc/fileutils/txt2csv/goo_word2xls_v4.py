import openpyxl
from openpyxl.styles import Alignment

def parse_vocabulary_to_excel(txt_filename, excel_filename):
    """
    解析单词本文件并导入到 Excel 文件中，处理长释义和分列显示。

    Args:
        txt_filename (str): 单词本文件名（.txt）。
        excel_filename (str): 输出 Excel 文件名（.xlsx）。
    """
    try:
        with open(txt_filename, 'r', encoding='utf-8') as f:
            content = f.read().strip()
    except FileNotFoundError:
        print(f"错误：找不到文件 '{txt_filename}'")
        return

    entries = content.split('\n\n')
    data = []
    for entry in entries:
        lines = [line.strip() for line in entry.strip().split('\n') if line.strip()]
        if len(lines) >= 2:
            word = lines[0]
            phonetic = lines[1]
            definitions = "\n".join(lines[2:])
            data.append({'word': word, 'phonetic': phonetic, 'definitions': definitions})

    num_words = len(data)
    half_num = (num_words + 1) // 2

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "词汇表"

    # 写入数据
    for i, entry in enumerate(data):
        row_index = i + 1
        col_offset = 0
        if i >= half_num:
            col_offset = 3
            row_index = i - half_num + 1

        word = entry['word']
        phonetic = entry['phonetic']
        definitions = entry['definitions']
        definition_lines = definitions.split('\n')
        num_definition_lines = len(definition_lines)

        # 写入单词和音标（只在起始行写入）
        if num_definition_lines > 0:
            sheet.cell(row=row_index, column=1 + col_offset, value=word)
            sheet.cell(row=row_index, column=2 + col_offset, value=phonetic)

            # 写入释义
            for j, line in enumerate(definition_lines):
                sheet.cell(row=row_index + j, column=3 + col_offset, value=line)

            # 合并单词和音标单元格
            if num_definition_lines > 1:
                sheet.merge_cells(start_row=row_index, start_column=1 + col_offset,
                                  end_row=row_index + num_definition_lines - 1, end_column=1 + col_offset)
                sheet.merge_cells(start_row=row_index, start_column=2 + col_offset,
                                  end_row=row_index + num_definition_lines - 1, end_column=2 + col_offset)
                # 设置垂直对齐
                for row in range(row_index, row_index + num_definition_lines):
                    sheet.cell(row=row, column=1 + col_offset).alignment = Alignment(vertical='top')
                    sheet.cell(row=row, column=2 + col_offset).alignment = Alignment(vertical='top')

    try:
        workbook.save(excel_filename)
        print(f"成功将数据导入到 '{excel_filename}'")
    except Exception as e:
        print(f"保存 Excel 文件时发生错误：{e}")

if __name__ == "__main__":
    txt_file = r"goo_word2xls_v3_wd\词汇-01-自然地理.txt"  # 替换为您的单词本文件名
    excel_file = r"goo_word2xls_v3_wd\自然地理词汇.xlsx"  # 替换为您想要保存的 Excel 文件名
    parse_vocabulary_to_excel(txt_file, excel_file)